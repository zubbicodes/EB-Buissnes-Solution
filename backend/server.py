from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import re
import csv
import io
import asyncio
import uuid
import logging
import bcrypt
import jwt
import base64
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, BackgroundTasks
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from rapidfuzz import fuzz, process as rf_process, utils as rf_utils
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from google.auth.transport import requests as google_requests
from google.oauth2 import id_token as google_id_token


# ----- DB / Config -----
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ['JWT_SECRET']
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "").strip()
APP_ENV = os.environ.get("APP_ENV", os.environ.get("ENV", "development")).lower()
COOKIE_SECURE = (os.environ.get("COOKIE_SECURE") or ("true" if APP_ENV == "production" else "false")).lower() == "true"
ALLOW_DEFAULT_ADMIN = (os.environ.get("ALLOW_DEFAULT_ADMIN") or ("true" if APP_ENV != "production" else "false")).lower() == "true"
DEFAULT_ORG_NAME = os.environ.get("DEFAULT_ORG_NAME", "EB Business Solutions Limited")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ----- Auth utils -----
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=30), "type": "refresh"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie("access_token", access, httponly=True, secure=COOKIE_SECURE, samesite="lax", max_age=7 * 24 * 3600, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=COOKIE_SECURE, samesite="lax", max_age=30 * 24 * 3600, path="/")


async def ensure_user_workspace(user: Dict[str, Any]) -> Dict[str, Any]:
    if user.get("org_id") and user.get("role"):
        return user
    now = datetime.now(timezone.utc).isoformat()
    org_id = user.get("org_id") or str(uuid.uuid4())
    if not hasattr(db, "organizations"):
        user["org_id"] = org_id
        user["role"] = user.get("role") or "admin"
        return user
    if not await db.organizations.find_one({"id": org_id}):
        await db.organizations.insert_one({
            "id": org_id,
            "name": user.get("organization_name") or DEFAULT_ORG_NAME,
            "created_at": now,
            "created_by": user["id"],
        })
    role = user.get("role") or "admin"
    await db.users.update_one({"id": user["id"]}, {"$set": {"org_id": org_id, "role": role, "updated_at": now}})
    user["org_id"] = org_id
    user["role"] = role
    return user

async def get_current_user(request: Request) -> Dict[str, Any]:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return await ensure_user_workspace(user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


def require_role(current: Dict[str, Any], allowed: List[str]):
    if current.get("role") not in allowed:
        raise HTTPException(status_code=403, detail="You do not have permission to perform this action")


def org_query(current: Dict[str, Any], extra: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    q = {"org_id": current["org_id"]}
    if extra:
        q.update(extra)
    return q


# ----- Models -----
class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)

class LoginIn(BaseModel):
    email: EmailStr
    password: str

class GoogleAuthIn(BaseModel):
    credential: str = Field(min_length=1)

class UploadFileIn(BaseModel):
    name: str = ""
    content_type: str = ""
    data_base64: str = ""
    text: str = ""

class ColumnMapping(BaseModel):
    # bank
    bank_date: Optional[str] = None
    bank_reference: Optional[str] = None
    bank_amount: Optional[str] = None
    bank_payer: Optional[str] = None
    bank_account: Optional[str] = None
    bank_transaction_type: Optional[str] = None
    # invoice
    invoice_number: Optional[str] = None
    invoice_debtor: Optional[str] = None
    invoice_amount: Optional[str] = None
    invoice_date: Optional[str] = None
    invoice_outstanding: Optional[str] = None
    invoice_due_date: Optional[str] = None
    invoice_customer_reference: Optional[str] = None

class AllocationCreate(BaseModel):
    name: str
    period: str
    bank_csv: str = ""
    invoice_csv: str = ""
    bank_file: Optional[UploadFileIn] = None
    invoice_file: Optional[UploadFileIn] = None
    mapping: ColumnMapping
    proceed_with_warnings: bool = False

class ValidateIn(BaseModel):
    bank_csv: str = ""
    invoice_csv: str = ""
    bank_file: Optional[UploadFileIn] = None
    invoice_file: Optional[UploadFileIn] = None
    mapping: ColumnMapping

class ManualLinkIn(BaseModel):
    bank_row_id: str
    invoice_row_id: str
    amount: float

class ManualUnlinkIn(BaseModel):
    bank_row_id: str
    invoice_row_id: str
    amount: Optional[float] = None

class ManualReallocateIn(BaseModel):
    from_bank_row_id: str
    from_invoice_row_id: str
    to_bank_row_id: Optional[str] = None
    to_invoice_row_id: str
    amount: float

class ExceptionUpdateIn(BaseModel):
    notes: Optional[str] = None
    status: Optional[str] = Field(default=None, pattern=r"^(open|reviewed)$")

class WorkspaceUserUpdateIn(BaseModel):
    role: str = Field(pattern=r"^(admin|user|read_only)$")

# ----- CSV helpers -----
def parse_csv(text: str) -> List[Dict[str, str]]:
    if not text or not text.strip():
        return []
    # Strip BOM that some Excel exports prepend
    if text.startswith("\ufeff"):
        text = text[1:]
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for r in reader:
        rows.append({(k or "").strip(): (v or "").strip() for k, v in r.items() if k is not None})
    return rows


def parse_xlsx_bytes(data: bytes) -> List[Dict[str, str]]:
    if not data:
        return []
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    while headers and not headers[-1]:
        headers.pop()
    if not headers:
        return []
    rows: List[Dict[str, str]] = []
    for values in rows_iter:
        if values is None:
            continue
        vals = list(values[:len(headers)])
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = vals[i] if i < len(vals) else ""
            if isinstance(v, datetime):
                row[h] = v.date().isoformat()
            else:
                row[h] = "" if v is None else str(v).strip()
        rows.append(row)
    return rows


def parse_upload(text: str = "", file_payload: Optional[UploadFileIn] = None) -> List[Dict[str, str]]:
    if file_payload:
        name = (file_payload.name or "").lower()
        ctype = (file_payload.content_type or "").lower()
        if file_payload.text and not file_payload.data_base64:
            return parse_csv(file_payload.text)
        raw = file_payload.data_base64 or ""
        if "," in raw and raw.strip().startswith("data:"):
            raw = raw.split(",", 1)[1]
        data = base64.b64decode(raw) if raw else b""
        if name.endswith(".xlsx") or "spreadsheetml" in ctype or "excel" in ctype:
            return parse_xlsx_bytes(data)
        try:
            return parse_csv(data.decode("utf-8-sig"))
        except UnicodeDecodeError:
            return parse_csv(data.decode("latin-1"))
    return parse_csv(text)


def headers_for_rows(rows: List[Dict[str, str]]) -> List[str]:
    return list(rows[0].keys()) if rows else []

def to_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("£", "").replace("$", "").replace("€", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None

def parse_date_safe(v) -> Optional[str]:
    if not v:
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def validate_rows(bank_rows: List[Dict[str, str]], inv_rows: List[Dict[str, str]], mapping: ColumnMapping) -> Dict[str, Any]:
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    coverage: Dict[str, float] = {}

    if not bank_rows:
        errors.append({"scope": "bank", "message": "Bank CSV is empty or unreadable."})
    if not inv_rows:
        errors.append({"scope": "invoice", "message": "Invoice CSV is empty or unreadable."})

    def check_col(rows, col, scope, key, required=True):
        if not col:
            if required:
                errors.append({"scope": scope, "message": f"Required column not mapped: {key}"})
            return
        if rows and col not in rows[0]:
            errors.append({"scope": scope, "message": f"Column '{col}' not found in {scope} CSV header."})
            return
        non_empty = sum(1 for r in rows if r.get(col, "").strip())
        cov = non_empty / len(rows) if rows else 0
        coverage[f"{scope}.{key}"] = round(cov * 100, 1)
        if cov < 0.5:
            warnings.append({"scope": scope, "message": f"Low coverage on {key} ({col}): {round(cov*100,1)}%"})

    check_col(bank_rows, mapping.bank_reference, "bank", "reference")
    check_col(bank_rows, mapping.bank_amount, "bank", "amount")
    check_col(bank_rows, mapping.bank_date, "bank", "date", required=False)
    check_col(bank_rows, mapping.bank_payer, "bank", "payer", required=False)
    check_col(bank_rows, mapping.bank_account, "bank", "account", required=False)
    check_col(bank_rows, mapping.bank_transaction_type, "bank", "transaction_type", required=False)

    check_col(inv_rows, mapping.invoice_number, "invoice", "number")
    check_col(inv_rows, mapping.invoice_debtor, "invoice", "debtor")
    check_col(inv_rows, mapping.invoice_amount, "invoice", "amount")
    check_col(inv_rows, mapping.invoice_outstanding, "invoice", "outstanding", required=False)
    check_col(inv_rows, mapping.invoice_date, "invoice", "date", required=False)
    check_col(inv_rows, mapping.invoice_due_date, "invoice", "due_date", required=False)
    check_col(inv_rows, mapping.invoice_customer_reference, "invoice", "customer_reference", required=False)

    # Row-level checks
    if mapping.bank_amount:
        zeros = 0
        negs = 0
        unparseable = 0
        for i, r in enumerate(bank_rows, start=2):
            amt = to_float(r.get(mapping.bank_amount, ""))
            if amt is None:
                unparseable += 1
                if unparseable <= 5:
                    warnings.append({"scope": "bank", "row": i, "message": f"Amount could not be parsed: '{r.get(mapping.bank_amount, '')}'"})
            elif amt == 0:
                zeros += 1
            elif amt < 0:
                negs += 1
        if bank_rows and zeros == len(bank_rows):
            errors.append({"scope": "bank", "message": "All bank amounts are zero."})
        if negs:
            warnings.append({"scope": "bank", "message": f"{negs} bank row(s) have negative amounts (refunds/withdrawals)."})

    # Invalid date format warnings
    def check_dates(rows, col, scope):
        if not col:
            return
        bad = 0
        for i, r in enumerate(rows, start=2):
            raw = r.get(col, "").strip()
            if raw and parse_date_safe(raw) is None:
                bad += 1
                if bad <= 3:
                    warnings.append({"scope": scope, "row": i, "message": f"Unrecognised date format in '{col}': '{raw}'"})
        if bad > 3:
            warnings.append({"scope": scope, "message": f"{bad} row(s) total have unparseable dates in '{col}'."})
    check_dates(bank_rows, mapping.bank_date, "bank")
    check_dates(inv_rows, mapping.invoice_date, "invoice")
    check_dates(inv_rows, mapping.invoice_due_date, "invoice")

    # Missing debtor names
    if mapping.invoice_debtor and inv_rows:
        missing_debtor = sum(1 for r in inv_rows if not r.get(mapping.invoice_debtor, "").strip())
        if missing_debtor:
            warnings.append({"scope": "invoice", "message": f"{missing_debtor} invoice row(s) have no debtor name — fuzzy matching will skip these."})

    if mapping.invoice_number and inv_rows:
        seen = {}
        for i, r in enumerate(inv_rows, start=2):
            num = r.get(mapping.invoice_number, "").strip()
            if not num:
                continue
            seen.setdefault(num, []).append(i)
        dups = {k: v for k, v in seen.items() if len(v) > 1}
        if dups:
            for k, rows in list(dups.items())[:5]:
                warnings.append({"scope": "invoice", "message": f"Duplicate invoice number '{k}' on rows {rows}"})

    return {
        "ok": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "coverage": coverage,
        "bank_row_count": len(bank_rows),
        "invoice_row_count": len(inv_rows),
    }


def validate_csvs(bank_csv: str, invoice_csv: str, mapping: ColumnMapping) -> Dict[str, Any]:
    return validate_rows(parse_csv(bank_csv), parse_csv(invoice_csv), mapping)


# ----- Matching engine -----
INVOICE_REF_PATTERN = re.compile(r"\b([A-Z]{0,5}-?\d{3,10})\b", re.IGNORECASE)

def extract_refs(text: str) -> List[str]:
    if not text:
        return []
    return [m.group(1).upper().replace("-", "") for m in INVOICE_REF_PATTERN.finditer(text)]

def normalize_num(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())


CORP_SUFFIXES = {
    "ltd", "limited", "llc", "inc", "incorporated", "co", "company", "plc",
    "corp", "corporation", "gmbh", "sa", "ag", "bv", "nv", "srl", "spa", "llp", "lp", "kg",
}

STOP_TOKENS = {"the", "and", "of", "for", "to", "from", "payment", "transfer",
               "bacs", "chq", "ref", "inv", "invoice", "kref", "receipt", "deposit"}


def distinctive_tokens(s: str) -> List[str]:
    """Tokens of a debtor name that are useful for substring matching: >=4 chars, not corp suffix, not stop word."""
    if not s:
        return []
    return [t.lower() for t in re.findall(r"[A-Za-z0-9]+", s)
            if len(t) >= 4 and t.lower() not in CORP_SUFFIXES and t.lower() not in STOP_TOKENS]


def fmt_money(v: float) -> str:
    try:
        return f"£{float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v)


def strip_corp_suffix(s: str) -> str:
    """Remove generic corporate suffix tokens for tighter fuzzy comparison."""
    if not s:
        return ""
    tokens = re.findall(r"[A-Za-z0-9]+", s)
    keep = [t for t in tokens if t.lower() not in CORP_SUFFIXES]
    return " ".join(keep) if keep else s


def invoice_fifo_key(inv: Dict[str, Any]):
    date = inv.get("date") or "9999-12-31"
    return (date, inv.get("idx", 0))


def debtor_norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", strip_corp_suffix(s or "").lower())


def allocate_link(b: Dict[str, Any], inv: Dict[str, Any], method: str, confidence: str,
                  reason: str, score: Optional[float] = None, ambiguous: bool = False,
                  ref_kind: Optional[str] = None) -> Optional[Dict[str, Any]]:
    if b["remaining"] <= 0 or inv["remaining"] <= 0:
        return None
    alloc = round(min(b["remaining"], inv["remaining"]), 2)
    if alloc <= 0:
        return None
    outstanding_before = inv["remaining"]
    b["remaining"] = round(b["remaining"] - alloc, 2)
    inv["remaining"] = round(inv["remaining"] - alloc, 2)
    link = {
        "bank_id": b["id"],
        "invoice_id": inv["id"],
        "amount": alloc,
        "method": method,
        "confidence": confidence,
        "reason": reason,
        "invoice_amount": inv["amount"],
        "invoice_outstanding_before": outstanding_before,
        "invoice_outstanding_after": inv["remaining"],
        "ambiguous": ambiguous,
    }
    if score is not None:
        link["score"] = round(score, 1)
    if ref_kind:
        link["ref_kind"] = ref_kind
    b["matches"].append(link)
    inv["matches"].append(link)
    return link


def run_matching(bank_rows_raw, invoice_rows_raw, mapping: ColumnMapping):
    # Build normalized rows
    bank_rows = []
    for i, r in enumerate(bank_rows_raw):
        amt = to_float(r.get(mapping.bank_amount or "", "")) or 0.0
        ref = r.get(mapping.bank_reference or "", "")
        payer = r.get(mapping.bank_payer or "", "") if mapping.bank_payer else ""
        date = parse_date_safe(r.get(mapping.bank_date or "", "")) if mapping.bank_date else None
        bank_rows.append({
            "id": str(uuid.uuid4()),
            "idx": i,
            "amount": round(amt, 2),
            "remaining": round(amt, 2),
            "reference": ref,
            "payer": payer,
            "date": date,
            "matches": [],
            "status": "unmatched",  # unmatched | partial | full
            "confidence": None,
            "extracted_refs": [],
            "best_debtor_score": None,
            "overpaid_amount": 0.0,
        })

    invoice_rows = []
    for i, r in enumerate(invoice_rows_raw):
        amt = to_float(r.get(mapping.invoice_amount or "", "")) or 0.0
        outstanding_field = mapping.invoice_outstanding
        outstanding = to_float(r.get(outstanding_field or "", "")) if outstanding_field else None
        if outstanding is None:
            outstanding = amt
        num = r.get(mapping.invoice_number or "", "")
        debtor = r.get(mapping.invoice_debtor or "", "")
        date = parse_date_safe(r.get(mapping.invoice_date or "", "")) if mapping.invoice_date else None
        invoice_rows.append({
            "id": str(uuid.uuid4()),
            "idx": i,
            "number": num,
            "number_norm": normalize_num(num),
            "debtor": debtor,
            "amount": round(amt, 2),
            "outstanding": round(outstanding, 2),
            "remaining": round(outstanding, 2),
            "date": date,
            "matches": [],
            "status": "unmatched",
        })

    inv_by_norm = {}
    inv_by_digits: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    inv_by_digit_suffix: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for inv in invoice_rows:
        if inv["number_norm"]:
            inv_by_norm.setdefault(inv["number_norm"], []).append(inv)
            digits = re.sub(r"\D", "", inv["number_norm"])
            if len(digits) >= 4:
                inv_by_digits[digits].append(inv)
                for suffix_len in range(4, len(digits) + 1):
                    inv_by_digit_suffix[digits[-suffix_len:]].append(inv)

    # Inverted token index: distinctive debtor token -> list of invoice indices.
    # Used to pre-filter candidates for Pass 2 + Pass 2.5 (huge speedup on large invoice sets).
    token_index: Dict[str, List[int]] = defaultdict(list)
    debtor_tokens_cache: List[List[str]] = []
    debtor_norm_cache: List[str] = []
    debtor_norm_index: Dict[str, List[int]] = defaultdict(list)
    for idx, inv in enumerate(invoice_rows):
        toks = distinctive_tokens(inv["debtor"]) if inv["debtor"] else []
        debtor_tokens_cache.append(toks)
        norm = debtor_norm(inv["debtor"]) if inv["debtor"] else ""
        debtor_norm_cache.append(norm)
        if norm:
            debtor_norm_index[norm].append(idx)
        for t in toks:
            token_index[t].append(idx)

    # Pass 1: invoice reference match.
    # CRITICAL HIERARCHY RULE: if Pass 1 finds ANY reference hits for a bank row,
    # Pass 2 (fuzzy debtor) and Pass 2.5 (token-substring) are SKIPPED for that row.
    # The remainder stays unallocated (PARTIAL) rather than being balanced by unrelated invoices.
    for b in bank_rows:
        if b["remaining"] <= 0:
            continue
        refs = extract_refs(b["reference"])
        b["extracted_refs"] = sorted(set(refs))  # stored for the review panel
        hit_invoices = []
        seen_ids = set()
        for ref in refs:
            cands = []
            if ref in inv_by_norm:
                cands = inv_by_norm[ref]
            else:
                digits = re.sub(r"\D", "", ref)
                if len(digits) >= 4:
                    cands.extend(inv_by_digit_suffix.get(digits, ()))
                    for suffix_len in range(4, len(digits) + 1):
                        cands.extend(inv_by_digits.get(digits[-suffix_len:], ()))
            for c in cands:
                if c["id"] in seen_ids:
                    continue
                seen_ids.add(c["id"])
                hit_invoices.append(c)

        hit_invoices.sort(key=invoice_fifo_key)
        for inv in hit_invoices:
            if b["remaining"] <= 0:
                break
            if inv["remaining"] <= 0:
                continue
            alloc = round(min(b["remaining"], inv["remaining"]), 2)
            if alloc <= 0:
                continue
            # Detect if the regex hit was exact vs partial (suffix-digit fallback)
            normalized_bank_ref = next(
                (r for r in refs if r == inv["number_norm"] or re.sub(r"\D", "", r).endswith(re.sub(r"\D", "", inv["number_norm"]))
                 or re.sub(r"\D", "", inv["number_norm"]).endswith(re.sub(r"\D", "", r))),
                None,
            )
            ref_kind = "exact" if normalized_bank_ref == inv["number_norm"] else "partial"
            outstanding_before = inv["remaining"]
            b["remaining"] = round(b["remaining"] - alloc, 2)
            inv["remaining"] = round(inv["remaining"] - alloc, 2)
            link = {
                "bank_id": b["id"],
                "invoice_id": inv["id"],
                "amount": alloc,
                "method": "reference",
                "ref_kind": ref_kind,
                "confidence": "high" if ref_kind == "exact" else "medium",
                "reason": (
                    f"Exact invoice reference detected ('{inv['number']}')"
                    if ref_kind == "exact"
                    else f"Partial invoice reference detected ('{inv['number']}' suffix match)"
                ),
                "invoice_amount": inv["amount"],
                "invoice_outstanding_before": outstanding_before,
                "invoice_outstanding_after": inv["remaining"],
            }
            b["matches"].append(link)
            inv["matches"].append(link)

        if hit_invoices and b["remaining"] > 0.005:
            b["overpaid_amount"] = b["remaining"]

    # Pass 1.5: exact debtor/customer name match. RUNS ONLY IF Pass 1 didn't match anything.
    for b in bank_rows:
        if b["remaining"] <= 0:
            continue
        if any(m["method"] == "reference" for m in b["matches"]):
            continue
        compare_text = " ".join([t for t in [b.get("reference") or "", b.get("payer") or ""] if t]).strip()
        compare_norm = debtor_norm(compare_text)
        if not compare_norm:
            continue
        cand_idx_set = set()
        for token in distinctive_tokens(compare_text):
            cand_idx_set.update(token_index.get(token, ()))
        exact_candidates = [
            invoice_rows[i] for i in cand_idx_set
            if invoice_rows[i]["remaining"] > 0
            and debtor_norm_cache[i]
            and debtor_norm_cache[i] in compare_norm
        ]
        if not exact_candidates:
            continue
        exact_candidates.sort(key=invoice_fifo_key)
        debtor_label = exact_candidates[0]["debtor"]
        ambiguous = len({debtor_norm(x["debtor"]) for x in exact_candidates}) > 1
        for inv in exact_candidates:
            if b["remaining"] <= 0:
                break
            allocate_link(
                b, inv, "debtor_exact", "high",
                f"Exact debtor/customer name match for {debtor_label}; allocated FIFO by oldest invoice",
                score=100.0,
                ambiguous=ambiguous,
            )
        b["best_debtor_score"] = 100.0

    # Pass 2: fuzzy debtor name (WRatio, threshold 70). RUNS ONLY IF earlier passes didn't match anything for the row.
    for b in bank_rows:
        if b["remaining"] <= 0:
            continue
        if b["matches"]:
            # HIERARCHY: skip fuzzy if an earlier pass matched this bank row
            continue
        compare_text = " ".join([t for t in [b.get("reference") or "", b.get("payer") or ""] if t]).strip()
        compare_text_clean = strip_corp_suffix(compare_text)
        if not compare_text_clean:
            continue
        bank_tokens_set = set(distinctive_tokens(compare_text))
        cand_idx_set = set()
        for t in bank_tokens_set:
            cand_idx_set.update(token_index.get(t, ()))
        if not cand_idx_set:
            continue
        candidates = [(i, invoice_rows[i]) for i in cand_idx_set
                      if invoice_rows[i]["remaining"] > 0 and invoice_rows[i]["debtor"]]
        if not candidates:
            continue
        debtor_cleans = [strip_corp_suffix(c["debtor"]) for _, c in candidates]
        matches = rf_process.extract(
            compare_text_clean,
            debtor_cleans,
            scorer=fuzz.WRatio,
            processor=rf_utils.default_process,
            score_cutoff=70,
            limit=10,
        )
        scored = []
        for _, score, ci in matches:
            scored.append((score, candidates[ci][1]))
        scored.sort(key=lambda x: (-x[0], invoice_fifo_key(x[1])))
        plausible_count = len(scored)
        if scored:
            b["best_debtor_score"] = round(scored[0][0], 1)
        if scored:
            best_score, best_inv = scored[0]
            best_norm = debtor_norm(best_inv["debtor"])
            scored = [
                (best_score, invoice_rows[i]) for i in debtor_norm_index.get(best_norm, ())
                if invoice_rows[i]["remaining"] > 0
            ]
            scored.sort(key=lambda x: invoice_fifo_key(x[1]))
        for score, inv in scored:
            if b["remaining"] <= 0:
                break
            if inv["remaining"] <= 0:
                continue
            alloc = round(min(b["remaining"], inv["remaining"]), 2)
            if alloc <= 0:
                continue
            outstanding_before = inv["remaining"]
            b["remaining"] = round(b["remaining"] - alloc, 2)
            inv["remaining"] = round(inv["remaining"] - alloc, 2)
            confidence = "high" if score >= 95 else "medium" if score >= 85 else "low"
            link = {
                "bank_id": b["id"],
                "invoice_id": inv["id"],
                "amount": alloc,
                "method": "debtor_name",
                "confidence": confidence,
                "score": round(score, 1),
                "ambiguous": plausible_count > 1,
                "reason": f"Debtor name similarity {round(score, 1)}%"
                          + (f" — {plausible_count} plausible candidates" if plausible_count > 1 else " — unique candidate"),
                "invoice_amount": inv["amount"],
                "invoice_outstanding_before": outstanding_before,
                "invoice_outstanding_after": inv["remaining"],
            }
            b["matches"].append(link)
            inv["matches"].append(link)

    # Pass 2.5: token-substring fallback. RUNS ONLY IF no Pass 1 or Pass 2 match for this row.
    for b in bank_rows:
        if b["remaining"] <= 0:
            continue
        if b["matches"]:
            # HIERARCHY: skip token-substring if any earlier pass matched
            continue
        bank_text_lc = " ".join([t for t in [b.get("reference") or "", b.get("payer") or ""] if t]).lower()
        if not bank_text_lc:
            continue
        bank_tokens_set = set(distinctive_tokens(bank_text_lc))
        if not bank_tokens_set:
            continue
        hit_counts: Dict[int, int] = defaultdict(int)
        for t in bank_tokens_set:
            for i in token_index.get(t, ()):
                hit_counts[i] += 1
        scored = []
        for i, hits in hit_counts.items():
            if hits < 2:
                continue
            inv = invoice_rows[i]
            if inv["remaining"] <= 0 or not inv["debtor"]:
                continue
            total_toks = len(debtor_tokens_cache[i])
            score = fuzz.partial_token_set_ratio(bank_text_lc, inv["debtor"].lower())
            scored.append((score, hits, total_toks, inv))
        scored.sort(key=lambda x: (-x[1], -x[0], -x[3]["remaining"]))
        plausible_count = len(scored)
        for score, hits, total_toks, inv in scored:
            if b["remaining"] <= 0:
                break
            alloc = round(min(b["remaining"], inv["remaining"]), 2)
            if alloc <= 0:
                continue
            outstanding_before = inv["remaining"]
            b["remaining"] = round(b["remaining"] - alloc, 2)
            inv["remaining"] = round(inv["remaining"] - alloc, 2)
            link = {
                "bank_id": b["id"],
                "invoice_id": inv["id"],
                "amount": alloc,
                "method": "debtor_tokens",
                "confidence": "low",
                "score": round(score, 1),
                "ambiguous": plausible_count > 1,
                "reason": f"Token-substring match ({hits}/{total_toks} distinctive debtor tokens found in bank text, low confidence — review required)"
                          + (f" — {plausible_count} plausible candidates" if plausible_count > 1 else ""),
                "invoice_amount": inv["amount"],
                "invoice_outstanding_before": outstanding_before,
                "invoice_outstanding_after": inv["remaining"],
            }
            b["matches"].append(link)
            inv["matches"].append(link)

    # ---- Final classification (finance-grade strict rules) ----
    for b in bank_rows:
        ms = b["matches"]
        if not ms:
            b["status"] = "unmatched"
            b["confidence"] = None
            b["reason"] = "No invoice reference detected and no debtor-name match above threshold"
            continue

        ref_methods = [m for m in ms if m["method"] == "reference"]
        debtor_name_methods = [m for m in ms if m["method"] in ("debtor_name", "debtor_exact")]
        debtor_methods = [m for m in ms if m["method"] in ("debtor_name", "debtor_exact", "debtor_tokens")]
        fully_consumed = b["remaining"] <= 0.005

        # Under strict hierarchy, Pass 2/2.5 never runs when Pass 1 has matches, so
        # ref_methods and debtor_methods are now mutually exclusive on the same bank row.

        # Rule A: reference match(es), fully consumed
        rule_a = bool(ref_methods) and fully_consumed
        # Rule B: debtor/customer match (NOT debtor_tokens), fully consumed, and not ambiguous.
        rule_b = (
            not ref_methods
            and debtor_name_methods
            and all(m.get("method") != "debtor_tokens" for m in ms)
            and not any(m.get("ambiguous") for m in debtor_name_methods)
            and max((m.get("score", 0) for m in debtor_name_methods), default=0) >= 85
            and fully_consumed
        )

        if b.get("overpaid_amount", 0) > 0.005:
            b["status"] = "overpaid"
            b["reason"] = f"Invoice reference matched; payment exceeds referenced invoice(s) by {fmt_money(b['overpaid_amount'])}"
            b["confidence"] = "high"
        elif rule_a or rule_b:
            b["status"] = "full"
            reason_parts = []
            if rule_a:
                # Pull invoice numbers from match links (now standardised after the reason wording change)
                inv_nums = sorted({m.get("invoice_number") for m in ref_methods if m.get("invoice_number")})
                refs = ", ".join(inv_nums) if inv_nums else ""
                reason_parts.append(
                    f"Invoice reference{'s' if len(ref_methods) > 1 else ''}"
                    + (f" {refs}" if refs else "")
                    + " matched and bank amount fully consumed"
                )
            if rule_b:
                m = debtor_name_methods[0]
                reason_parts.append(f"Unique debtor match at {m['score']}% with exact amount")
            b["reason"] = " · ".join(reason_parts)
            b["confidence"] = "high"
        else:
            b["status"] = "partial"
            # Highest confidence among matches, capped to "medium" unless reference-based
            conf_order = {"high": 3, "medium": 2, "low": 1}
            best = max((m["confidence"] for m in ms), key=lambda c: conf_order.get(c, 0))
            if not ref_methods and best == "high":
                # Demote debtor-only matches with score 95+ but ambiguous or partial-amount
                best = "medium"
            b["confidence"] = best
            ambiguous = any(m.get("ambiguous") for m in debtor_methods)
            bits = []
            if ref_methods and not fully_consumed:
                bits.append(f"Reference match but {fmt_money(b['remaining'])} unallocated")
            if debtor_methods:
                m_best = max(debtor_methods, key=lambda x: x.get("score", 0))
                if m_best.get("method") == "debtor_tokens":
                    bits.append(f"Token-substring match {m_best.get('score', '?')}% (low confidence — review required)")
                else:
                    bits.append(f"Debtor name similarity {m_best.get('score', '?')}%")
                if ambiguous:
                    bits.append("multiple candidates — needs review")
            if not ref_methods and not debtor_methods:
                bits.append("partial reference allocation")
            b["reason"] = "; ".join(bits) or "Requires review"

    for inv in invoice_rows:
        if not inv["matches"]:
            inv["status"] = "unmatched"
        elif inv["remaining"] <= 0.005:
            inv["status"] = "full"
        else:
            inv["status"] = "partial"

    stats = {
        "total_bank": len(bank_rows),
        "total_invoices": len(invoice_rows),
        "fully_matched": sum(1 for b in bank_rows if b["status"] == "full"),
        "partially_matched": sum(1 for b in bank_rows if b["status"] == "partial"),
        "overpaid": sum(1 for b in bank_rows if b["status"] == "overpaid"),
        "manual": sum(1 for b in bank_rows if any(m.get("method") == "manual" for m in b.get("matches", []))),
        "unmatched_bank": sum(1 for b in bank_rows if b["status"] == "unmatched"),
        "unmatched_invoices": sum(1 for inv in invoice_rows if inv["status"] == "unmatched"),
        "total_allocated": round(sum(m["amount"] for b in bank_rows for m in b["matches"]), 2),
        "total_outstanding": round(sum(inv["remaining"] for inv in invoice_rows), 2),
    }
    return bank_rows, invoice_rows, stats


# ----- App -----
app = FastAPI()
api = APIRouter(prefix="/api")


@api.get("/")
async def root():
    return {"app": "Receivables Reconciliation Platform", "company": "EB Business Solutions Limited"}


# ----- Auth routes -----
@api.post("/auth/register")
async def register(payload: RegisterIn, response: Response):
    email = payload.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user_id = str(uuid.uuid4())
    org_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    await db.organizations.insert_one({
        "id": org_id,
        "name": payload.name or DEFAULT_ORG_NAME,
        "created_at": now,
        "created_by": user_id,
    })
    doc = {
        "id": user_id,
        "email": email,
        "name": payload.name,
        "password_hash": hash_password(payload.password),
        "org_id": org_id,
        "role": "admin",
        "created_at": now,
    }
    await db.users.insert_one(doc)
    access = create_access_token(user_id, email)
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)
    return {"id": user_id, "email": email, "name": payload.name, "org_id": org_id, "role": "admin", "access_token": access, "refresh_token": refresh}


@api.post("/auth/login")
async def login(payload: LoginIn, response: Response):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not user.get("password_hash") or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    now = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"id": user["id"]}, {"$set": {"last_login_at": now, "updated_at": now}})
    access = create_access_token(user["id"], email)
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    user = await ensure_user_workspace(user)
    return {"id": user["id"], "email": user["email"], "name": user.get("name", ""), "org_id": user["org_id"], "role": user["role"], "access_token": access, "refresh_token": refresh}


@api.post("/auth/google")
async def google_auth(payload: GoogleAuthIn, response: Response):
    if not GOOGLE_CLIENT_ID:
        raise HTTPException(status_code=503, detail="Google Sign-In is not configured")

    try:
        google_payload = google_id_token.verify_oauth2_token(
            payload.credential,
            google_requests.Request(),
            GOOGLE_CLIENT_ID,
        )
    except ValueError:
        raise HTTPException(status_code=401, detail="Invalid Google credential")

    email = (google_payload.get("email") or "").strip().lower()
    if not email:
        raise HTTPException(status_code=401, detail="Google account did not provide an email")
    if not google_payload.get("email_verified"):
        raise HTTPException(status_code=401, detail="Google email is not verified")

    google_sub = google_payload.get("sub")
    now = datetime.now(timezone.utc).isoformat()
    name = (google_payload.get("name") or email.split("@")[0]).strip()
    picture = google_payload.get("picture")

    user = await db.users.find_one({"email": email})
    if user:
        user_id = user["id"]
        user = await ensure_user_workspace(user)
        update = {
            "auth_provider": "google",
            "google_sub": google_sub,
            "picture": picture,
            "last_login_at": now,
            "updated_at": now,
        }
        if name and not user.get("name"):
            update["name"] = name
        await db.users.update_one({"id": user_id}, {"$set": update})
        user_name = update.get("name") or user.get("name", "") or name
        org_id = user["org_id"]
        role = user["role"]
    else:
        user_id = str(uuid.uuid4())
        org_id = str(uuid.uuid4())
        user_name = name
        if hasattr(db, "organizations"):
            await db.organizations.insert_one({
                "id": org_id,
                "name": user_name or DEFAULT_ORG_NAME,
                "created_at": now,
                "created_by": user_id,
            })
        await db.users.insert_one({
            "id": user_id,
            "email": email,
            "name": user_name,
            "org_id": org_id,
            "role": "admin",
            "auth_provider": "google",
            "google_sub": google_sub,
            "picture": picture,
            "created_at": now,
            "last_login_at": now,
            "updated_at": now,
        })
        role = "admin"

    access = create_access_token(user_id, email)
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)
    return {"id": user_id, "email": email, "name": user_name, "org_id": org_id, "role": role, "access_token": access, "refresh_token": refresh}


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"ok": True}


@api.post("/auth/refresh")
async def refresh_session(request: Request, response: Response):
    refresh = request.cookies.get("refresh_token")
    if not refresh:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(refresh, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")
    access = create_access_token(user["id"], user["email"])
    new_refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, new_refresh)
    return user


@api.get("/auth/me")
async def me(current=Depends(get_current_user)):
    return current


# ----- CSV column-mapping presets (built-in) and saved profiles (per-user) -----
BUILT_IN_PRESETS = [
    {
        "id": "barclays_business",
        "label": "Barclays Business (UK)",
        "scope": "bank",
        "mapping": {"bank_date": "Date", "bank_reference": "Memo", "bank_payer": "Subcategory", "bank_amount": "Amount", "bank_account": "Account Number"},
    },
    {
        "id": "hsbc_uk",
        "label": "HSBC UK",
        "scope": "bank",
        "mapping": {"bank_date": "Date", "bank_reference": "Reference", "bank_payer": "Payee", "bank_amount": "Paid In", "bank_transaction_type": "Type"},
    },
    {
        "id": "lloyds_uk",
        "label": "Lloyds UK",
        "scope": "bank",
        "mapping": {"bank_date": "Transaction Date", "bank_reference": "Transaction Description", "bank_amount": "Credit Amount", "bank_transaction_type": "Transaction Type"},
    },
    {
        "id": "xero_aged_debtors",
        "label": "Xero Aged Receivables",
        "scope": "invoice",
        "mapping": {"invoice_number": "Invoice Number", "invoice_debtor": "Customer", "invoice_amount": "Invoice Amount", "invoice_outstanding": "Outstanding", "invoice_date": "Invoice Date", "invoice_due_date": "Due Date"},
    },
    {
        "id": "sage_aged_debtors",
        "label": "Sage Aged Debtors",
        "scope": "invoice",
        "mapping": {"invoice_number": "Reference", "invoice_debtor": "Account Name", "invoice_amount": "Net", "invoice_outstanding": "Balance", "invoice_date": "Date"},
    },
    {
        "id": "quickbooks_aged",
        "label": "QuickBooks A/R Aging",
        "scope": "invoice",
        "mapping": {"invoice_number": "Num", "invoice_debtor": "Customer", "invoice_amount": "Amount", "invoice_outstanding": "Open Balance", "invoice_date": "Date", "invoice_due_date": "Due Date"},
    },
]


@api.get("/mapping/presets")
async def list_presets(current=Depends(get_current_user)):
    user_profiles = await db.user_mapping_profiles.find(
        {"org_id": current["org_id"]}, {"_id": 0, "user_id": 0}
    ).sort("created_at", -1).to_list(100)
    return {"built_in": BUILT_IN_PRESETS, "saved": user_profiles}


class SaveMappingIn(BaseModel):
    label: str = Field(min_length=1, max_length=80)
    scope: str = Field(pattern=r"^(bank|invoice|both)$")
    mapping: Dict[str, Any]


@api.post("/mapping/presets")
async def save_preset(payload: SaveMappingIn, current=Depends(get_current_user)):
    require_role(current, ["admin", "user"])
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": current["id"],
        "org_id": current["org_id"],
        "label": payload.label,
        "scope": payload.scope,
        "mapping": payload.mapping,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.user_mapping_profiles.insert_one(doc)
    doc.pop("user_id", None)
    doc.pop("_id", None)
    return doc


@api.delete("/mapping/presets/{preset_id}")
async def delete_preset(preset_id: str, current=Depends(get_current_user)):
    require_role(current, ["admin", "user"])
    res = await db.user_mapping_profiles.delete_one({"id": preset_id, "org_id": current["org_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Profile not found")
    return {"ok": True}


# ----- Allocation routes -----
@api.post("/allocations/validate")
async def validate(payload: ValidateIn, current=Depends(get_current_user)):
    bank_rows = parse_upload(payload.bank_csv, payload.bank_file)
    inv_rows = parse_upload(payload.invoice_csv, payload.invoice_file)
    result = validate_rows(bank_rows, inv_rows, payload.mapping)
    await write_audit(
        current["id"],
        None,
        "validate_upload",
        {
            "ok": result["ok"],
            "bank_rows": result["bank_row_count"],
            "invoice_rows": result["invoice_row_count"],
            "errors": len(result["errors"]),
            "warnings": len(result["warnings"]),
        },
        current["org_id"],
    )
    return result


@api.post("/allocations/preview-headers")
async def preview_headers(payload: Dict[str, Any], current=Depends(get_current_user)):
    """Quick header sniffer for the wizard."""
    bank_file = UploadFileIn(**payload["bank_file"]) if payload.get("bank_file") else None
    invoice_file = UploadFileIn(**payload["invoice_file"]) if payload.get("invoice_file") else None
    bank_rows = parse_upload(payload.get("bank_csv", ""), bank_file)
    inv_rows = parse_upload(payload.get("invoice_csv", ""), invoice_file)
    return {
        "bank_headers": headers_for_rows(bank_rows),
        "bank_sample": bank_rows[:5],
        "invoice_headers": headers_for_rows(inv_rows),
        "invoice_sample": inv_rows[:5],
    }


async def write_audit(user_id: str, run_id: Optional[str], action: str, details: Dict[str, Any], org_id: Optional[str] = None):
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "org_id": org_id,
        "run_id": run_id,
        "action": action,
        "details": details,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })


# ----- Split-collection storage helpers -----
# For large runs (e.g. 29k+ invoice rows) we cannot embed everything in the allocation_runs doc
# because of the 16 MB BSON limit. Bank/invoice rows live in their own collections, scoped by run_id.

CHUNK = 1000  # bulk insert chunk size

async def _bulk_insert_rows(collection, run_id: str, user_id: str, org_id: str, rows: List[Dict[str, Any]]):
    if not rows:
        return
    for start in range(0, len(rows), CHUNK):
        batch = [{**r, "run_id": run_id, "user_id": user_id, "org_id": org_id} for r in rows[start:start + CHUNK]]
        await collection.insert_many(batch, ordered=False)


async def save_run_rows(run_id: str, user_id: str, org_id: str, bank_rows: List[Dict[str, Any]], invoice_rows: List[Dict[str, Any]]):
    await _bulk_insert_rows(db.allocation_bank_rows, run_id, user_id, org_id, bank_rows)
    await _bulk_insert_rows(db.allocation_invoice_rows, run_id, user_id, org_id, invoice_rows)


async def load_all_bank_rows(run_id: str, org_id: str) -> List[Dict[str, Any]]:
    return await db.allocation_bank_rows.find(
        {"run_id": run_id, "org_id": org_id}, {"_id": 0, "run_id": 0, "user_id": 0, "org_id": 0}
    ).sort("idx", 1).to_list(None)


async def load_all_invoice_rows(run_id: str, org_id: str) -> List[Dict[str, Any]]:
    return await db.allocation_invoice_rows.find(
        {"run_id": run_id, "org_id": org_id}, {"_id": 0, "run_id": 0, "user_id": 0, "org_id": 0}
    ).sort("idx", 1).to_list(None)


async def delete_run_rows(run_id: str, org_id: str):
    await db.allocation_bank_rows.delete_many({"run_id": run_id, "org_id": org_id})
    await db.allocation_invoice_rows.delete_many({"run_id": run_id, "org_id": org_id})
    await db.exceptions.delete_many({"run_id": run_id, "org_id": org_id})


def enrich_matches(bank_rows: List[Dict[str, Any]], invoice_rows: List[Dict[str, Any]]):
    """Denormalise invoice_number / invoice_debtor into each match for self-contained display + paginated reads."""
    inv_by_id = {inv["id"]: inv for inv in invoice_rows}
    for b in bank_rows:
        for m in b.get("matches", []):
            inv = inv_by_id.get(m.get("invoice_id"))
            if inv:
                m["invoice_number"] = inv.get("number")
                m["invoice_debtor"] = inv.get("debtor")
    for inv in invoice_rows:
        for m in inv.get("matches", []):
            # noop; bank denormalisation not strictly needed for our UI
            pass


async def rebuild_exceptions(run_id: str, user_id: str, org_id: str):
    await db.exceptions.delete_many({"run_id": run_id, "org_id": org_id})
    now = datetime.now(timezone.utc).isoformat()
    docs: List[Dict[str, Any]] = []
    bank_rows = await load_all_bank_rows(run_id, org_id)
    invoice_rows = await load_all_invoice_rows(run_id, org_id)

    seen_payments: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for b in bank_rows:
        key = f"{(b.get('date') or '').strip()}|{(b.get('reference') or '').strip().lower()}|{round(float(b.get('amount') or 0), 2)}"
        if b.get("reference"):
            seen_payments[key].append(b)
        if b.get("status") == "unmatched":
            docs.append({
                "id": str(uuid.uuid4()), "org_id": org_id, "user_id": user_id, "run_id": run_id,
                "type": "unmatched_payment", "status": "open", "bank_row_id": b["id"],
                "debtor": b.get("payer"), "reference": b.get("reference"), "amount": b.get("amount"),
                "confidence": b.get("confidence"), "notes": "", "created_at": now,
            })
        if b.get("status") == "partial":
            docs.append({
                "id": str(uuid.uuid4()), "org_id": org_id, "user_id": user_id, "run_id": run_id,
                "type": "underpayment", "status": "open", "bank_row_id": b["id"],
                "debtor": b.get("payer"), "reference": b.get("reference"), "amount": b.get("remaining"),
                "confidence": b.get("confidence"), "notes": "", "created_at": now,
            })
        if b.get("status") == "overpaid":
            docs.append({
                "id": str(uuid.uuid4()), "org_id": org_id, "user_id": user_id, "run_id": run_id,
                "type": "overpayment", "status": "open", "bank_row_id": b["id"],
                "debtor": b.get("payer"), "reference": b.get("reference"), "amount": b.get("overpaid_amount") or b.get("remaining"),
                "confidence": b.get("confidence"), "notes": "", "created_at": now,
            })
        if b.get("confidence") == "low" or any(m.get("confidence") == "low" for m in b.get("matches", [])):
            docs.append({
                "id": str(uuid.uuid4()), "org_id": org_id, "user_id": user_id, "run_id": run_id,
                "type": "low_confidence", "status": "open", "bank_row_id": b["id"],
                "debtor": b.get("payer"), "reference": b.get("reference"), "amount": b.get("amount"),
                "confidence": b.get("confidence") or "low", "notes": "", "created_at": now,
            })

    for rows in seen_payments.values():
        if len(rows) > 1:
            for b in rows:
                docs.append({
                    "id": str(uuid.uuid4()), "org_id": org_id, "user_id": user_id, "run_id": run_id,
                    "type": "duplicate_payment", "status": "open", "bank_row_id": b["id"],
                    "debtor": b.get("payer"), "reference": b.get("reference"), "amount": b.get("amount"),
                    "confidence": b.get("confidence"), "notes": "", "created_at": now,
                })

    seen_invoices: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for inv in invoice_rows:
        if inv.get("number"):
            seen_invoices[inv["number"].strip().lower()].append(inv)
        if inv.get("status") == "unmatched":
            docs.append({
                "id": str(uuid.uuid4()), "org_id": org_id, "user_id": user_id, "run_id": run_id,
                "type": "unmatched_invoice", "status": "open", "invoice_row_id": inv["id"],
                "debtor": inv.get("debtor"), "invoice_number": inv.get("number"), "amount": inv.get("remaining"),
                "confidence": None, "notes": "", "created_at": now,
            })
    for rows in seen_invoices.values():
        if len(rows) > 1:
            for inv in rows:
                docs.append({
                    "id": str(uuid.uuid4()), "org_id": org_id, "user_id": user_id, "run_id": run_id,
                    "type": "duplicate_invoice", "status": "open", "invoice_row_id": inv["id"],
                    "debtor": inv.get("debtor"), "invoice_number": inv.get("number"), "amount": inv.get("amount"),
                    "confidence": None, "notes": "", "created_at": now,
                })

    if docs:
        await db.exceptions.insert_many(docs, ordered=False)
    await db.allocation_runs.update_one(
        {"id": run_id, "org_id": org_id},
        {"$set": {"stats.exceptions": len(docs)}},
    )
    return len(docs)


async def _process_run_async(run_id: str, user_id: str, org_id: str, bank_raw: List[Dict[str, str]], inv_raw: List[Dict[str, str]], mapping_dict: Dict[str, Any]):
    """Background processor for large allocation runs."""
    try:
        await db.allocation_runs.update_one(
            {"id": run_id, "org_id": org_id},
            {"$set": {"progress": 10, "progress_phase": "Preparing and matching rows"}},
        )
        mapping = ColumnMapping(**mapping_dict)
        # Matching is CPU-heavy. Run it off the asyncio event loop so users,
        # audit, progress polling, and all other API endpoints remain responsive.
        bank_rows, invoice_rows, stats = await asyncio.to_thread(
            run_matching, bank_raw, inv_raw, mapping
        )
        await db.allocation_runs.update_one(
            {"id": run_id, "org_id": org_id},
            {"$set": {"progress": 60, "progress_phase": "Enriching match results"}},
        )
        enrich_matches(bank_rows, invoice_rows)
        await db.allocation_runs.update_one(
            {"id": run_id, "org_id": org_id},
            {"$set": {"progress": 75, "progress_phase": "Saving allocation results"}},
        )
        await save_run_rows(run_id, user_id, org_id, bank_rows, invoice_rows)
        await db.allocation_runs.update_one(
            {"id": run_id, "org_id": org_id},
            {"$set": {"progress": 90, "progress_phase": "Building exception report"}},
        )
        stats["exceptions"] = await rebuild_exceptions(run_id, user_id, org_id)
        await db.allocation_runs.update_one(
            {"id": run_id, "org_id": org_id},
            {"$set": {
                "stats": stats,
                "status": "done",
                "progress": 100,
                "progress_phase": "Allocation complete",
                "completed_at": datetime.now(timezone.utc).isoformat(),
            }},
        )
        await write_audit(user_id, run_id, "create_run", {"stats": stats, "async": True}, org_id)
    except Exception as e:
        logger.exception("Async run failed: %s", e)
        await db.allocation_runs.update_one(
            {"id": run_id, "org_id": org_id},
            {"$set": {"status": "error", "error": str(e), "progress_phase": "Allocation failed"}},
        )
        await write_audit(user_id, run_id, "create_run_failed", {"error": str(e)}, org_id)


@api.post("/allocations")
async def create_allocation(payload: AllocationCreate, background: BackgroundTasks, current=Depends(get_current_user)):
    require_role(current, ["admin", "user"])
    bank_raw = parse_upload(payload.bank_csv, payload.bank_file)
    inv_raw = parse_upload(payload.invoice_csv, payload.invoice_file)
    validation = validate_rows(bank_raw, inv_raw, payload.mapping)
    if not validation["ok"]:
        raise HTTPException(status_code=400, detail={"message": "CSV validation failed", "validation": validation})
    if validation["warnings"] and not payload.proceed_with_warnings:
        raise HTTPException(status_code=400, detail={"message": "CSV has warnings", "validation": validation})

    bank_row_count = validation.get("bank_row_count", 0)
    invoice_row_count = validation.get("invoice_row_count", 0)

    # Anything > 2000 on either side runs in the background to keep HTTP fast
    is_large = bank_row_count > 2000 or invoice_row_count > 2000

    run_id = str(uuid.uuid4())
    base_doc = {
        "id": run_id,
        "user_id": current["id"],
        "org_id": current["org_id"],
        "created_by": current["id"],
        "name": payload.name,
        "period": payload.period,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "mapping": payload.mapping.model_dump(),
        "status": "processing" if is_large else "done",
        "progress": 5 if is_large else 100,
        "progress_phase": "Queued for processing" if is_large else "Allocation complete",
    }

    if is_large:
        await db.allocation_runs.insert_one({**base_doc, "stats": {
            "total_bank": bank_row_count, "total_invoices": invoice_row_count,
            "fully_matched": 0, "partially_matched": 0, "unmatched_bank": 0,
            "unmatched_invoices": 0, "overpaid": 0, "manual": 0, "exceptions": 0,
            "total_allocated": 0.0, "total_outstanding": 0.0,
        }})
        await write_audit(
            current["id"],
            run_id,
            "create_run_queued",
            {"name": payload.name, "period": payload.period, "bank_rows": bank_row_count, "invoice_rows": invoice_row_count},
            current["org_id"],
        )
        background.add_task(_process_run_async, run_id, current["id"], current["org_id"], bank_raw, inv_raw, payload.mapping.model_dump())
        doc = await db.allocation_runs.find_one({"id": run_id, "org_id": current["org_id"]}, {"_id": 0})
        return doc

    bank_rows, invoice_rows, stats = run_matching(bank_raw, inv_raw, payload.mapping)
    enrich_matches(bank_rows, invoice_rows)

    doc = {**base_doc, "stats": stats}
    await db.allocation_runs.insert_one(doc)
    await save_run_rows(run_id, current["id"], current["org_id"], bank_rows, invoice_rows)
    stats["exceptions"] = await rebuild_exceptions(run_id, current["id"], current["org_id"])
    await db.allocation_runs.update_one({"id": run_id, "org_id": current["org_id"]}, {"$set": {"stats": stats}})
    doc["stats"] = stats
    await write_audit(current["id"], run_id, "create_run", {
        "name": payload.name, "period": payload.period, "stats": stats,
    }, current["org_id"])
    doc.pop("_id", None)
    return doc


@api.get("/allocations")
async def list_allocations(current=Depends(get_current_user)):
    runs = await db.allocation_runs.find(
        {"org_id": current["org_id"], "status": {"$ne": "archived"}},
        {"_id": 0, "bank_rows": 0, "invoice_rows": 0},
    ).sort("created_at", -1).to_list(500)
    return runs


@api.get("/allocations/{run_id}")
async def get_allocation(run_id: str, current=Depends(get_current_user)):
    """Returns the run header (stats + mapping + status), NOT the full row data.
    Use /allocations/{id}/rows for paginated row data."""
    run = await db.allocation_runs.find_one(
        {"id": run_id, "org_id": current["org_id"]},
        {"_id": 0, "bank_rows": 0, "invoice_rows": 0},
    )
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    return run


@api.get("/allocations/{run_id}/rows")
async def get_allocation_rows(
    run_id: str,
    bucket: str = "full",
    page: int = 1,
    page_size: int = 50,
    search: str = "",
    current=Depends(get_current_user),
):
    run = await db.allocation_runs.find_one({"id": run_id, "org_id": current["org_id"]}, {"_id": 0, "id": 1})
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    page = max(1, page)
    page_size = max(1, min(500, page_size))
    skip = (page - 1) * page_size

    if bucket in ("full", "partial", "overpaid", "unmatched_bank"):
        status_map = {"full": "full", "partial": "partial", "overpaid": "overpaid", "unmatched_bank": "unmatched"}
        q: Dict[str, Any] = {"run_id": run_id, "org_id": current["org_id"], "status": status_map[bucket]}
        if search:
            esc = re.escape(search)
            q["$or"] = [
                {"reference": {"$regex": esc, "$options": "i"}},
                {"payer": {"$regex": esc, "$options": "i"}},
                {"matches.invoice_number": {"$regex": esc, "$options": "i"}},
                {"matches.invoice_debtor": {"$regex": esc, "$options": "i"}},
            ]
        total = await db.allocation_bank_rows.count_documents(q)
        rows = await db.allocation_bank_rows.find(
            q, {"_id": 0, "run_id": 0, "user_id": 0, "org_id": 0}
        ).sort("idx", 1).skip(skip).limit(page_size).to_list(page_size)
    elif bucket == "unmatched_invoice":
        q = {"run_id": run_id, "org_id": current["org_id"], "status": "unmatched"}
        if search:
            esc = re.escape(search)
            q["$or"] = [
                {"number": {"$regex": esc, "$options": "i"}},
                {"debtor": {"$regex": esc, "$options": "i"}},
            ]
        total = await db.allocation_invoice_rows.count_documents(q)
        rows = await db.allocation_invoice_rows.find(
            q, {"_id": 0, "run_id": 0, "user_id": 0, "org_id": 0}
        ).sort("idx", 1).skip(skip).limit(page_size).to_list(page_size)
    else:
        raise HTTPException(status_code=400, detail=f"Unknown bucket: {bucket}")

    return {"rows": rows, "page": page, "page_size": page_size, "total": total, "bucket": bucket}


@api.delete("/allocations/{run_id}")
async def delete_allocation(run_id: str, current=Depends(get_current_user)):
    require_role(current, ["admin"])
    run = await db.allocation_runs.find_one({"id": run_id, "org_id": current["org_id"]}, {"_id": 0, "name": 1, "created_at": 1})
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    bank_result = await db.allocation_bank_rows.delete_many({"run_id": run_id, "org_id": current["org_id"]})
    invoice_result = await db.allocation_invoice_rows.delete_many({"run_id": run_id, "org_id": current["org_id"]})
    exception_result = await db.exceptions.delete_many({"run_id": run_id, "org_id": current["org_id"]})
    await db.allocation_runs.delete_one({"id": run_id, "org_id": current["org_id"]})
    await write_audit(
        current["id"],
        run_id,
        "delete_run",
        {
            "name": run.get("name"),
            "bank_rows_deleted": bank_result.deleted_count,
            "invoice_rows_deleted": invoice_result.deleted_count,
            "exceptions_deleted": exception_result.deleted_count,
        },
        current["org_id"],
    )
    return {"ok": True, "deleted": True}


@api.get("/allocations/{run_id}/export")
async def export_allocation(run_id: str, current=Depends(get_current_user)):
    run = await db.allocation_runs.find_one({"id": run_id, "org_id": current["org_id"]}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    async def stream():
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(["Bank Date", "Bank Reference", "Bank Payer", "Bank Amount (£)",
                         "Status", "Confidence", "Why matched?", "Matched Invoices", "Allocated (£)", "Bank Remaining (£)"])
        yield out.getvalue()
        async for b in db.allocation_bank_rows.find(
            {"run_id": run_id, "org_id": current["org_id"]}, {"_id": 0}
        ).sort("idx", 1):
            out = io.StringIO()
            writer = csv.writer(out)
            nums = ", ".join((m.get("invoice_number") or "?") for m in b.get("matches", []))
            allocated = round(sum(m["amount"] for m in b.get("matches", [])), 2)
            writer.writerow([
                b.get("date") or "",
                b.get("reference") or "",
                b.get("payer") or "",
                f"{b['amount']:.2f}",
                b["status"],
                b.get("confidence") or "",
                b.get("reason") or "",
                nums,
                f"{allocated:.2f}",
                f"{b['remaining']:.2f}",
            ])
            yield out.getvalue()

    return StreamingResponse(
        stream(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="allocation-{run["name"]}.csv"'},
    )


@api.get("/allocations/{run_id}/export-xlsx")
async def export_allocation_xlsx(run_id: str, current=Depends(get_current_user)):
    """Stream a large Excel export without OOMing.

    Uses openpyxl's write-only mode (~10x lower memory, ~5x faster) so 29k+ rows
    don't blow the worker. Per-row colour fills are skipped — only the header is
    styled — because per-cell PatternFill allocations are the dominant cost on
    large workbooks. The final BytesIO is yielded in 64 KB chunks so the client
    receives bytes incrementally instead of buffering the whole file.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.cell import WriteOnlyCell
    run = await db.allocation_runs.find_one({"id": run_id, "org_id": current["org_id"]}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    wb = Workbook(write_only=True)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="0F172A")
    hdr_align = Alignment(horizontal="left")

    def _hdr(ws_ref, value):
        c = WriteOnlyCell(ws_ref, value=value)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        return c

    ws = wb.create_sheet("Allocations")
    ws.append([_hdr(ws, h) for h in [
        "Bank Date", "Bank Reference", "Bank Payer", "Bank Amount", "Status",
        "Confidence", "Why matched?", "Matched Invoices", "Allocated", "Bank Remaining",
    ]])

    async for b in db.allocation_bank_rows.find(
        {"run_id": run_id, "org_id": current["org_id"]}, {"_id": 0}
    ).sort("idx", 1):
        nums = ", ".join((m.get("invoice_number") or "?") for m in b.get("matches", []))
        allocated = round(sum(m["amount"] for m in b.get("matches", [])), 2)
        ws.append([
            b.get("date") or "",
            b.get("reference") or "",
            b.get("payer") or "",
            b["amount"],
            b["status"],
            b.get("confidence") or "",
            b.get("reason") or "",
            nums,
            allocated,
            b["remaining"],
        ])

    ws2 = wb.create_sheet("Invoices")
    ws2.append([_hdr(ws2, h) for h in [
        "Invoice #", "Debtor", "Date", "Amount", "Outstanding", "Status",
    ]])
    async for inv in db.allocation_invoice_rows.find(
        {"run_id": run_id, "org_id": current["org_id"]}, {"_id": 0}
    ).sort("idx", 1):
        ws2.append([
            inv["number"],
            inv.get("debtor") or "",
            inv.get("date") or "",
            inv["amount"],
            inv["remaining"],
            inv["status"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    size = buf.getbuffer().nbytes

    def chunked():
        CHUNK_BYTES = 64 * 1024
        while True:
            data = buf.read(CHUNK_BYTES)
            if not data:
                break
            yield data

    safe_name = re.sub(r"[^A-Za-z0-9_.-]", "_", run["name"])[:80]
    return StreamingResponse(
        chunked(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="allocation-{safe_name}.xlsx"',
            "Content-Length": str(size),
        },
    )


@api.post("/allocations/{run_id}/manual-link")
async def manual_link(run_id: str, payload: ManualLinkIn, current=Depends(get_current_user)):
    require_role(current, ["admin", "user"])
    run = await db.allocation_runs.find_one({"id": run_id, "org_id": current["org_id"]}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    bank = await db.allocation_bank_rows.find_one(
        {"run_id": run_id, "org_id": current["org_id"], "id": payload.bank_row_id}, {"_id": 0, "run_id": 0, "user_id": 0, "org_id": 0}
    )
    inv = await db.allocation_invoice_rows.find_one(
        {"run_id": run_id, "org_id": current["org_id"], "id": payload.invoice_row_id}, {"_id": 0, "run_id": 0, "user_id": 0, "org_id": 0}
    )
    if not bank or not inv:
        raise HTTPException(status_code=404, detail="Bank or invoice row not found")
    amt = round(min(payload.amount, bank["remaining"], inv["remaining"]), 2)
    if amt <= 0:
        raise HTTPException(status_code=400, detail="Nothing remaining to allocate")

    link = {
        "bank_id": bank["id"], "invoice_id": inv["id"], "amount": amt,
        "method": "manual", "confidence": "manual",
        "reason": "Manual allocation",
        "invoice_number": inv.get("number"),
        "invoice_debtor": inv.get("debtor"),
        "invoice_amount": inv.get("amount"),
        "invoice_outstanding_before": inv["remaining"],
        "invoice_outstanding_after": round(inv["remaining"] - amt, 2),
    }
    bank.setdefault("matches", []).append(link)
    inv.setdefault("matches", []).append(link)
    bank["remaining"] = round(bank["remaining"] - amt, 2)
    inv["remaining"] = round(inv["remaining"] - amt, 2)
    bank["status"] = "manual" if bank["remaining"] <= 0.005 else "partial"
    bank["confidence"] = "manual"
    inv["status"] = "full" if inv["remaining"] <= 0.005 else "partial"

    await db.allocation_bank_rows.update_one(
        {"run_id": run_id, "org_id": current["org_id"], "id": bank["id"]},
        {"$set": {"matches": bank["matches"], "remaining": bank["remaining"],
                  "status": bank["status"], "confidence": bank["confidence"], "reason": "Includes manual override"}},
    )
    await db.allocation_invoice_rows.update_one(
        {"run_id": run_id, "org_id": current["org_id"], "id": inv["id"]},
        {"$set": {"matches": inv["matches"], "remaining": inv["remaining"], "status": inv["status"]}},
    )

    # Recompute stats from the source of truth (the split collections)
    await rebuild_exceptions(run_id, current["id"], current["org_id"])
    stats = await _recompute_stats(run_id, current["org_id"])
    await db.allocation_runs.update_one(
        {"id": run_id, "org_id": current["org_id"]},
        {"$set": {"stats": stats}},
    )
    await write_audit(current["id"], run_id, "manual_link", {
        "bank_reference": bank.get("reference"),
        "invoice_number": inv.get("number"),
        "amount": amt,
    }, current["org_id"])
    return {"ok": True, "stats": stats, "bank": bank, "invoice": inv}


@api.post("/allocations/{run_id}/manual-unlink")
async def manual_unlink(run_id: str, payload: ManualUnlinkIn, current=Depends(get_current_user)):
    require_role(current, ["admin", "user"])
    run = await db.allocation_runs.find_one({"id": run_id, "org_id": current["org_id"]}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    bank = await db.allocation_bank_rows.find_one({"run_id": run_id, "org_id": current["org_id"], "id": payload.bank_row_id}, {"_id": 0})
    inv = await db.allocation_invoice_rows.find_one({"run_id": run_id, "org_id": current["org_id"], "id": payload.invoice_row_id}, {"_id": 0})
    if not bank or not inv:
        raise HTTPException(status_code=404, detail="Bank or invoice row not found")
    matches = bank.get("matches", [])
    target = next((m for m in matches if m.get("invoice_id") == payload.invoice_row_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="Allocation link not found")
    amt = round(min(float(payload.amount or target["amount"]), float(target["amount"])), 2)
    remaining_link = round(float(target["amount"]) - amt, 2)
    if remaining_link > 0.005:
        target["amount"] = remaining_link
    else:
        bank["matches"] = [m for m in matches if m is not target]
    inv["matches"] = [
        ({**m, "amount": remaining_link} if m.get("bank_id") == bank["id"] and remaining_link > 0.005 else m)
        for m in inv.get("matches", [])
        if not (m.get("bank_id") == bank["id"] and remaining_link <= 0.005)
    ]
    bank["remaining"] = round(float(bank.get("remaining") or 0) + amt, 2)
    inv["remaining"] = round(float(inv.get("remaining") or 0) + amt, 2)
    bank["status"] = "unmatched" if not bank["matches"] else "partial"
    inv["status"] = "unmatched" if not inv["matches"] else ("full" if inv["remaining"] <= 0.005 else "partial")
    await db.allocation_bank_rows.update_one(
        {"run_id": run_id, "org_id": current["org_id"], "id": bank["id"]},
        {"$set": {"matches": bank["matches"], "remaining": bank["remaining"], "status": bank["status"], "reason": "Manual allocation unlinked"}},
    )
    await db.allocation_invoice_rows.update_one(
        {"run_id": run_id, "org_id": current["org_id"], "id": inv["id"]},
        {"$set": {"matches": inv["matches"], "remaining": inv["remaining"], "status": inv["status"]}},
    )
    await rebuild_exceptions(run_id, current["id"], current["org_id"])
    stats = await _recompute_stats(run_id, current["org_id"])
    await db.allocation_runs.update_one({"id": run_id, "org_id": current["org_id"]}, {"$set": {"stats": stats}})
    await write_audit(current["id"], run_id, "manual_unlink", {
        "bank_reference": bank.get("reference"), "invoice_number": inv.get("number"), "amount": amt,
    }, current["org_id"])
    return {"ok": True, "stats": stats}


@api.post("/allocations/{run_id}/manual-reallocate")
async def manual_reallocate(run_id: str, payload: ManualReallocateIn, current=Depends(get_current_user)):
    await manual_unlink(run_id, ManualUnlinkIn(
        bank_row_id=payload.from_bank_row_id,
        invoice_row_id=payload.from_invoice_row_id,
        amount=payload.amount,
    ), current)
    result = await manual_link(run_id, ManualLinkIn(
        bank_row_id=payload.to_bank_row_id or payload.from_bank_row_id,
        invoice_row_id=payload.to_invoice_row_id,
        amount=payload.amount,
    ), current)
    await write_audit(current["id"], run_id, "manual_reallocate", payload.model_dump(), current["org_id"])
    return result


async def _recompute_stats(run_id: str, org_id: str) -> Dict[str, Any]:
    q = {"run_id": run_id, "org_id": org_id}
    total_bank = await db.allocation_bank_rows.count_documents(q)
    total_invoices = await db.allocation_invoice_rows.count_documents(q)
    full = await db.allocation_bank_rows.count_documents({**q, "status": "full"})
    partial = await db.allocation_bank_rows.count_documents({**q, "status": "partial"})
    unmatched_bank = await db.allocation_bank_rows.count_documents({**q, "status": "unmatched"})
    overpaid = await db.allocation_bank_rows.count_documents({**q, "status": "overpaid"})
    manual = await db.allocation_bank_rows.count_documents({**q, "status": "manual"})
    unmatched_inv = await db.allocation_invoice_rows.count_documents({**q, "status": "unmatched"})
    exceptions = await db.exceptions.count_documents(q)
    agg_alloc = await db.allocation_bank_rows.aggregate([
        {"$match": q},
        {"$unwind": {"path": "$matches", "preserveNullAndEmptyArrays": False}},
        {"$group": {"_id": None, "total": {"$sum": "$matches.amount"}}},
    ]).to_list(1)
    total_allocated = round(agg_alloc[0]["total"], 2) if agg_alloc else 0.0
    agg_out = await db.allocation_invoice_rows.aggregate([
        {"$match": q},
        {"$group": {"_id": None, "total": {"$sum": "$remaining"}}},
    ]).to_list(1)
    total_outstanding = round(agg_out[0]["total"], 2) if agg_out else 0.0
    return {
        "total_bank": total_bank, "total_invoices": total_invoices,
        "fully_matched": full, "partially_matched": partial,
        "overpaid": overpaid, "manual": manual,
        "unmatched_bank": unmatched_bank, "unmatched_invoices": unmatched_inv,
        "exceptions": exceptions, "total_allocated": total_allocated, "total_outstanding": total_outstanding,
    }


# ----- Compare -----
@api.get("/compare")
async def compare(run_ids: str, current=Depends(get_current_user)):
    ids = [i for i in run_ids.split(",") if i]
    if not ids:
        raise HTTPException(status_code=400, detail="No run_ids provided")
    runs = await db.allocation_runs.find(
        {"id": {"$in": ids}, "org_id": current["org_id"]},
        {"_id": 0, "id": 1, "name": 1, "period": 1, "created_at": 1},
    ).to_list(50)
    runs_by_id = {r["id"]: r for r in runs}
    ordered = [runs_by_id[i] for i in ids if i in runs_by_id]

    matrix: Dict[str, Dict[str, Any]] = {}
    for r in ordered:
        async for inv in db.allocation_invoice_rows.find(
            {"run_id": r["id"], "org_id": current["org_id"]},
            {"_id": 0, "debtor": 1, "remaining": 1, "status": 1, "number": 1},
        ):
            debtor = (inv.get("debtor") or "Unknown").strip() or "Unknown"
            row = matrix.setdefault(debtor, {"debtor": debtor, "runs": {}})
            entry = row["runs"].get(r["id"], {"status": "absent", "outstanding": 0, "invoices": []})
            entry["status"] = inv["status"]
            entry["outstanding"] = round(entry["outstanding"] + inv["remaining"], 2)
            entry["invoices"].append(inv["number"])
            row["runs"][r["id"]] = entry

    def unmatched_count(row):
        return sum(1 for r in ordered if row["runs"].get(r["id"], {}).get("status") == "unmatched")

    rows = list(matrix.values())
    rows.sort(key=lambda x: -unmatched_count(x))
    consistently_unmatched = [
        row["debtor"] for row in rows
        if all(row["runs"].get(r["id"], {}).get("status") == "unmatched" for r in ordered)
    ]
    return {
        "runs": [{"id": r["id"], "name": r["name"], "period": r["period"], "created_at": r["created_at"]} for r in ordered],
        "rows": rows,
        "consistently_unmatched": consistently_unmatched,
    }


# ----- Debtor report -----
@api.get("/debtors")
async def debtors(threshold: float = 0.0, current=Depends(get_current_user)):
    runs = await db.allocation_runs.find(
        {"org_id": current["org_id"]},
        {"_id": 0, "id": 1, "name": 1, "period": 1},
    ).sort("created_at", -1).to_list(500)
    agg: Dict[str, Dict[str, Any]] = {}
    for r in runs:
        async for inv in db.allocation_invoice_rows.find(
            {"run_id": r["id"], "org_id": current["org_id"]},
            {"_id": 0, "debtor": 1, "remaining": 1, "matches": 1, "status": 1, "number": 1},
        ):
            debtor = (inv.get("debtor") or "Unknown").strip() or "Unknown"
            d = agg.setdefault(debtor, {
                "debtor": debtor, "total_outstanding": 0.0, "total_allocated": 0.0,
                "invoice_count": 0, "runs": [],
            })
            d["total_outstanding"] = round(d["total_outstanding"] + inv["remaining"], 2)
            allocated = round(sum(m["amount"] for m in inv.get("matches", [])), 2)
            d["total_allocated"] = round(d["total_allocated"] + allocated, 2)
            d["invoice_count"] += 1
            d["runs"].append({
                "run_id": r["id"], "run_name": r["name"], "period": r["period"],
                "status": inv["status"], "outstanding": inv["remaining"], "allocated": allocated,
                "invoice_number": inv["number"],
            })
    rows = list(agg.values())
    for row in rows:
        row["flagged"] = bool(threshold) and row["total_outstanding"] >= threshold
    rows.sort(key=lambda x: (-int(x["flagged"]), -x["total_outstanding"]))
    return {
        "threshold": threshold,
        "rows": rows,
        "total_debtors": len(rows),
        "flagged_count": sum(1 for r in rows if r["flagged"]),
        "total_outstanding": round(sum(r["total_outstanding"] for r in rows), 2),
        "total_allocated": round(sum(r["total_allocated"] for r in rows), 2),
    }


@api.get("/debtors/export")
async def debtors_export(threshold: float = 0.0, current=Depends(get_current_user)):
    report = await debtors(threshold, current)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Debtor", "Flagged", "Total Outstanding (£)", "Total Allocated (£)", "Invoice Count", "Runs Appeared"])
    for r in report["rows"]:
        writer.writerow([
            r["debtor"], "YES" if r["flagged"] else "",
            f"{r['total_outstanding']:.2f}", f"{r['total_allocated']:.2f}",
            r["invoice_count"], len(r["runs"]),
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.read()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="debtor-report.csv"'},
    )


# ----- Exceptions -----
@api.get("/exceptions")
async def list_exceptions(
    run_id: Optional[str] = None,
    type: Optional[str] = None,
    status: Optional[str] = None,
    debtor: Optional[str] = None,
    current=Depends(get_current_user),
):
    q: Dict[str, Any] = {"org_id": current["org_id"]}
    if run_id:
        q["run_id"] = run_id
    if type:
        q["type"] = type
    if status:
        q["status"] = status
    if debtor:
        q["debtor"] = {"$regex": re.escape(debtor), "$options": "i"}
    rows = await db.exceptions.find(q, {"_id": 0, "org_id": 0}).sort("created_at", -1).to_list(1000)
    summary = {
        "total": len(rows),
        "open": sum(1 for r in rows if r.get("status") == "open"),
        "reviewed": sum(1 for r in rows if r.get("status") == "reviewed"),
        "by_type": {},
    }
    for r in rows:
        summary["by_type"][r["type"]] = summary["by_type"].get(r["type"], 0) + 1
    return {"summary": summary, "rows": rows}


@api.patch("/exceptions/{exception_id}")
async def update_exception(exception_id: str, payload: ExceptionUpdateIn, current=Depends(get_current_user)):
    require_role(current, ["admin", "user"])
    exc = await db.exceptions.find_one({"id": exception_id, "org_id": current["org_id"]}, {"_id": 0})
    if not exc:
        raise HTTPException(status_code=404, detail="Exception not found")
    update: Dict[str, Any] = {"updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current["id"]}
    if payload.notes is not None:
        update["notes"] = payload.notes
    if payload.status:
        update["status"] = payload.status
        if payload.status == "reviewed":
            update["reviewed_at"] = datetime.now(timezone.utc).isoformat()
            update["reviewed_by"] = current["id"]
        else:
            update["reviewed_at"] = None
            update["reviewed_by"] = None
    await db.exceptions.update_one({"id": exception_id, "org_id": current["org_id"]}, {"$set": update})
    await write_audit(current["id"], exc.get("run_id"), "exception_update", {"exception_id": exception_id, **update}, current["org_id"])
    updated = await db.exceptions.find_one({"id": exception_id, "org_id": current["org_id"]}, {"_id": 0, "org_id": 0})
    return updated


@api.get("/exceptions/export")
async def export_exceptions(run_id: Optional[str] = None, current=Depends(get_current_user)):
    data = await list_exceptions(run_id=run_id, current=current)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Type", "Status", "Run ID", "Debtor", "Reference", "Invoice", "Amount", "Confidence", "Notes", "Created", "Reviewed"])
    for r in data["rows"]:
        writer.writerow([
            r.get("type"), r.get("status"), r.get("run_id"), r.get("debtor") or "",
            r.get("reference") or "", r.get("invoice_number") or "", r.get("amount") or "",
            r.get("confidence") or "", r.get("notes") or "", r.get("created_at") or "",
            r.get("reviewed_at") or "",
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.read()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="exception-report.csv"'},
    )


# ----- Audit -----
@api.get("/audit")
async def audit(run_id: Optional[str] = None, current=Depends(get_current_user)):
    q: Dict[str, Any] = {"org_id": current["org_id"]}
    if run_id:
        q["run_id"] = run_id
    logs = await db.audit_logs.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    summary = {
        "total": len(logs),
        "create_run": sum(1 for x in logs if x["action"] == "create_run"),
        "delete_run": sum(1 for x in logs if x["action"] in ("delete_run", "archive_run")),
        "manual_link": sum(1 for x in logs if x["action"] == "manual_link"),
    }
    return {"summary": summary, "logs": logs}


# ----- Workspace users -----
@api.get("/workspace/users")
async def workspace_users(current=Depends(get_current_user)):
    require_role(current, ["admin"])
    users = await db.users.find(
        {"org_id": current["org_id"]},
        {"_id": 0, "password_hash": 0, "google_sub": 0},
    ).sort("created_at", 1).to_list(200)
    org = await db.organizations.find_one({"id": current["org_id"]}, {"_id": 0})
    return {"organization": org, "users": users}


@api.patch("/workspace/users/{user_id}")
async def update_workspace_user(user_id: str, payload: WorkspaceUserUpdateIn, current=Depends(get_current_user)):
    require_role(current, ["admin"])
    if user_id == current["id"] and payload.role != "admin":
        admins = await db.users.count_documents({"org_id": current["org_id"], "role": "admin"})
        if admins <= 1:
            raise HTTPException(status_code=400, detail="At least one admin is required")
    res = await db.users.update_one(
        {"id": user_id, "org_id": current["org_id"]},
        {"$set": {"role": payload.role, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    await write_audit(current["id"], None, "user_role_update", {"user_id": user_id, "role": payload.role}, current["org_id"])
    user = await db.users.find_one({"id": user_id, "org_id": current["org_id"]}, {"_id": 0, "password_hash": 0, "google_sub": 0})
    return user


app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    if APP_ENV == "production" and (JWT_SECRET == "your-secret-key-change-in-production" or len(JWT_SECRET) < 32):
        raise RuntimeError("Production JWT_SECRET must be set to a strong secret of at least 32 characters")
    await db.users.create_index("email", unique=True)
    await db.organizations.create_index("id", unique=True)
    await db.users.create_index([("org_id", 1), ("role", 1)])
    await db.allocation_runs.create_index([("user_id", 1), ("created_at", -1)])
    await db.allocation_runs.create_index([("org_id", 1), ("created_at", -1)])
    await db.allocation_runs.create_index([("id", 1), ("user_id", 1)])
    await db.allocation_runs.create_index([("id", 1), ("org_id", 1)])
    await db.allocation_bank_rows.create_index([("run_id", 1), ("user_id", 1), ("idx", 1)])
    await db.allocation_bank_rows.create_index([("run_id", 1), ("user_id", 1), ("status", 1)])
    await db.allocation_bank_rows.create_index([("run_id", 1), ("user_id", 1), ("id", 1)])
    await db.allocation_bank_rows.create_index([("run_id", 1), ("org_id", 1), ("status", 1)])
    await db.allocation_invoice_rows.create_index([("run_id", 1), ("user_id", 1), ("idx", 1)])
    await db.allocation_invoice_rows.create_index([("run_id", 1), ("user_id", 1), ("status", 1)])
    await db.allocation_invoice_rows.create_index([("run_id", 1), ("user_id", 1), ("id", 1)])
    await db.allocation_invoice_rows.create_index([("run_id", 1), ("org_id", 1), ("status", 1)])
    await db.audit_logs.create_index([("user_id", 1), ("created_at", -1)])
    await db.audit_logs.create_index([("org_id", 1), ("created_at", -1)])
    await db.exceptions.create_index([("run_id", 1), ("org_id", 1), ("type", 1), ("status", 1)])
    default_org = await db.organizations.find_one({"name": DEFAULT_ORG_NAME})
    if not default_org:
        default_org = {"id": str(uuid.uuid4()), "name": DEFAULT_ORG_NAME, "created_at": datetime.now(timezone.utc).isoformat()}
        await db.organizations.insert_one(default_org)
    default_org_id = default_org["id"]
    await db.users.update_many({"org_id": {"$exists": False}}, {"$set": {"org_id": default_org_id, "role": "admin"}})
    await db.allocation_runs.update_many({"org_id": {"$exists": False}}, {"$set": {"org_id": default_org_id}})
    await db.allocation_bank_rows.update_many({"org_id": {"$exists": False}}, {"$set": {"org_id": default_org_id}})
    await db.allocation_invoice_rows.update_many({"org_id": {"$exists": False}}, {"$set": {"org_id": default_org_id}})
    await db.audit_logs.update_many({"org_id": {"$exists": False}}, {"$set": {"org_id": default_org_id}})
    await db.user_mapping_profiles.update_many({"org_id": {"$exists": False}}, {"$set": {"org_id": default_org_id}})
    # Seed admin
    admin_email = (os.environ.get("ADMIN_EMAIL") or "admin@ebbusiness.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD") or "Admin@2026!"
    existing = await db.users.find_one({"email": admin_email})
    if not existing and ALLOW_DEFAULT_ADMIN:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "name": "EB Admin",
            "org_id": default_org_id,
            "role": "admin",
            "password_hash": hash_password(admin_password),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info(f"Seeded admin: {admin_email}")


@app.on_event("shutdown")
async def shutdown():
    client.close()
